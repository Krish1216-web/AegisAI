import pytest
import os
import tempfile
import openpyxl

from app.services.extractors.xlsx import XLSXExtractor

def test_xlsx_extractor_success():
    extractor = XLSXExtractor()

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as f:
        temp_path = f.name
        
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Overview"
        
        ws["A1"] = "Department"
        ws["B1"] = "Budget"
        ws["A2"] = "Engineering"
        ws["B2"] = 150000
        
        wb.save(temp_path)
        
        res = extractor.extract(temp_path)
        
        assert "Department | Budget" in res.text
        assert "Engineering | 150000" in res.text
        assert len(res.sections) == 1
        assert res.sections[0].title == "Overview"
        
    finally:
        if os.path.exists(temp_path):
            os.remove(temp_path)
