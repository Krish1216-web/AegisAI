import os
import openpyxl
from loguru import logger

from app.services.extractors.base import BaseDocumentExtractor, ExtractedDocument, SectionData
from app.core.document_exceptions import InvalidFile

MAX_ROWS_PER_SHEET = 1000
MAX_CELLS_PER_SHEET = 10000

class XLSXExtractor(BaseDocumentExtractor):
    def extract(self, file_path: str) -> ExtractedDocument:
        if not os.path.exists(file_path):
            raise InvalidFile("XLSX file does not exist on disk.")

        try:
            # Load with read_only and data_only to optimize speed and memory
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sections_data = []
            full_text_parts = []
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheet_rows = []
                
                cell_count = 0
                row_count = 0
                
                # Iterate rows safely up to safety limits
                for row in ws.iter_rows(values_only=True):
                    row_count += 1
                    if row_count > MAX_ROWS_PER_SHEET:
                        logger.warning(f"Sheet {sheet_name} exceeded max row limit {MAX_ROWS_PER_SHEET}. Truncating.")
                        sheet_rows.append("[Truncated: row limit reached]")
                        break
                        
                    # Filter trailing empty cells but keep internal columns aligned
                    non_empty = [str(cell) if cell is not None else "" for cell in row]
                    # If whole row is empty, skip it
                    if not any(non_empty):
                        continue
                        
                    cell_count += len(non_empty)
                    if cell_count > MAX_CELLS_PER_SHEET:
                        logger.warning(f"Sheet {sheet_name} exceeded max cell limit {MAX_CELLS_PER_SHEET}. Truncating.")
                        sheet_rows.append("[Truncated: cell limit reached]")
                        break
                        
                    row_str = " | ".join(non_empty)
                    sheet_rows.append(row_str)
                    
                sheet_text = "\n".join(sheet_rows)
                sections_data.append(
                    SectionData(
                        title=sheet_name,
                        text=sheet_text,
                        metadata={"sheet_name": sheet_name}
                    )
                )
                full_text_parts.append(f"### Sheet: {sheet_name}\n{sheet_text}")
                
            # Close workbook to free memory resource
            wb.close()

            full_text = "\n\n".join(full_text_parts).strip()
            char_count = len(full_text)
            word_count = len(full_text.split())

            return ExtractedDocument(
                text=full_text,
                pages=[],
                sections=sections_data,
                metadata={"sheet_names": wb.sheetnames},
                page_count=1,
                character_count=char_count,
                word_count=word_count
            )

        except Exception as e:
            logger.error(f"Failed to parse XLSX: {e}")
            raise InvalidFile(f"Failed to parse XLSX workbook: {str(e)}")
